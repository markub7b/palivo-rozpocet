from flask import Flask, render_template, request, redirect, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import numbers
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///nafta.db'
db = SQLAlchemy(app)

class Tankovanie(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    datum = db.Column(db.String(10), nullable=False)
    litre = db.Column(db.Float, nullable=False)
    cena_za_1l = db.Column(db.Float, nullable=False)
    typ_jazdy = db.Column(db.String(10), nullable=False)
    zaplatena_suma = db.Column(db.Float, nullable=False)
    celkova_cena = db.Column(db.Float)
    cena_na_osobu = db.Column(db.Float)

with app.app_context():
    db.create_all()

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        datum = request.form.get('datum')
        litre = float(request.form.get('litre'))
        cena_za_1l = float(request.form.get('cena_za_1l'))
        typ_jazdy = request.form.get('typ_jazdy')
        zaplatena_suma = float(request.form.get('zaplatena_suma'))

        celkova_cena = litre * cena_za_1l
        cena_na_osobu = None if typ_jazdy == 'súkromná' else celkova_cena / 4

        tankovanie = Tankovanie(datum=datum, litre=litre, cena_za_1l=cena_za_1l, typ_jazdy=typ_jazdy, zaplatena_suma=zaplatena_suma, celkova_cena=celkova_cena, cena_na_osobu=cena_na_osobu)
        db.session.add(tankovanie)
        db.session.commit()

        return redirect(url_for('result'))

    return render_template('index.html')

@app.route('/result')
def result():
    tankovania = Tankovanie.query.all()
    return render_template('result.html', data=tankovania)

@app.route('/delete/<int:id>', methods=['GET'])
def delete(id):
    tankovanie_to_delete = Tankovanie.query.get_or_404(id)
    db.session.delete(tankovanie_to_delete)
    db.session.commit()
    return redirect(url_for('result'))

@app.route('/download', methods=['GET'])
def download():
    tankovania = Tankovanie.query.all()
    data = {
        'Datum': [t.datum for t in tankovania],
        'Litre': [t.litre for t in tankovania],
        'Cena za 1L': [t.cena_za_1l for t in tankovania],
        'Typ jazdy': [t.typ_jazdy for t in tankovania],
        'Zaplatena suma': [t.zaplatena_suma for t in tankovania],
        'Cena na osobu': [t.cena_na_osobu for t in tankovania]
    }

    df = pd.DataFrame(data)
    file_path = 'tankovanie.xlsx'

    # Vytvoríme workbook pomocou openpyxl
    wb = Workbook()
    ws = wb.active

    # Pridáme dáta z DataFrame do worksheetu
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Nastavíme formátovanie stĺpca pre ceny
    price_column = ['Cena za 1L', 'Zaplatena suma', 'Cena na osobu']
    for cell in ws[1]:
        if cell.value in price_column:
            cell.number_format = u'€ #,##0.00'

    # Uložíme workbook
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)

@app.route('/download_selected', methods=['GET'])
def download_selected():
    selected_ids = request.args.get('ids').split(',')
    tankovania = Tankovanie.query.filter(Tankovanie.id.in_(selected_ids)).all()
    data = {
        'Datum': [t.datum for t in tankovania],
        'Litre': [t.litre for t in tankovania],
        'Cena za 1L': [t.cena_za_1l for t in tankovania],
        'Typ jazdy': [t.typ_jazdy for t in tankovania],
        'Zaplatena suma': [t.zaplatena_suma for t in tankovania],
        'Cena na osobu': [t.cena_na_osobu for t in tankovania]
    }

    df = pd.DataFrame(data)
    file_path = 'tankovanie_selected.xlsx'

    # Vytvoríme workbook pomocou openpyxl
    wb = Workbook()
    ws = wb.active

    # Pridáme dáta z DataFrame do worksheetu
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Nastavíme formátovanie stĺpca pre ceny
    price_column = ['Cena za 1L', 'Zaplatena suma', 'Cena na osobu']
    for cell in ws[1]:
        if cell.value in price_column:
            cell.number_format = u'€ #,##0.00'

    # Uložíme workbook
    wb.save(file_path)

    return send_file(file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True)
